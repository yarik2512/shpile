# Скрипт преобразования template.xlsx файла и импорт данных в базу данных
from openpyxl import load_workbook
from string import ascii_letters, digits
from random import sample
from mysql.connector import connect, Error

con = connect(
    host='37.140.192.174',
    database='u1490660_default',
    user='u1490660_default',
    password='Ds3Nb2d5wYj6UW28'
)
cur = con.cursor()

cur.execute(
    f"SELECT * FROM `groups`"
)
gr = cur.fetchall()
print(gr)
id_groups = {}
for item in gr:
    id_groups[item[1]] = item[0]

wb = load_workbook('import_student_template.xlsx')
ws = wb[wb.sheetnames[0]]
row = 2
while not (ws[f"A{row}"].value is None):
    email = ws[f"A{row}"].value
    fio = ''.join([ws[f"C{row}"].value, ws[f"D{row}"].value, ws[f"E{row}"].value])
    groups = [id_groups[item.value] for item in ws[f"F{row}:ZZ{row}"][0] if not (item.value is None)]
    groups = ';'.join(map(str, groups))
    password = ''.join(sample(ascii_letters + digits, 8))
    ws[f"B{row}"] = password
    cur.execute(
        f"INSERT INTO students "
        f"VALUES (NULL, '{email}', '{fio}', '{password}', '{groups}')"
    )

    row += 1
con.commit()
wb.save('import_student_with_password.xlsx')
