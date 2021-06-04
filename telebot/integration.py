import sqlite3
import csv
from openpyxl import load_workbook
import functions


def import_curriculum_new(con, cur, files):
    """Функция import_curriculum импортирует таблицы с учебными планами в БД.
        - аргументы: con - подключение к БД, cur - курсор БД, files - список имён файлов с учебными планами;
        - значение - нет."""
    for file in files:
        # загружаем учебный план
        wb = load_workbook(filename=file)
        year = int(file.split('.')[0][3:])
        # получаем список листов в книге
        sheets = wb.sheetnames
        # проходим по листам
        for sheet in sheets:
            # узнаем id направления подготовки
            abb, id_sub_direction = sheet.split('-')
            cur.execute(
                f"SELECT id FROM config WHERE abb={abb} and id_subdirection={id_sub_direction} and year={year}"
            )
            direction = cur.fetchone()[0]

            # тут ещё много нужно чего написать


def import_curriculum(cur, data):
    """Функция import_curriculum_to_db импортирует в БД словарь, представляющий собой учебный план."""
    for direct in data:
        direction, sub_direction, year = direct.split('-')[0], int(direct.split('-')[1]), int(direct.split('-')[2])

        for i in range(len(data[direct])):
            current_subject = data[direct][i]
            for subject in current_subject:
                level, burden = current_subject[subject]
                half_10_1, report_10_1, weeks_10_1, half_10_2, report_10_2, weeks_10_2 = burden['101'][0], \
                                                                                         burden['101'][1], \
                                                                                         burden['101'][2], \
                                                                                         burden['102'][0], \
                                                                                         burden['102'][1], \
                                                                                         burden['102'][2]
                half_10_1 = half_10_1 if not (half_10_1 is None) else 0
                half_10_2 = half_10_2 if not (half_10_2 is None) else 0

                n_year_10 = weeks_10_1 * half_10_1 + weeks_10_2 * half_10_2

                report_10_1 = report_10_1 if not (report_10_1 is None) else 'NULL'
                report_10_2 = report_10_2 if not (report_10_2 is None) else 'NULL'
                cur.execute(
                    f"INSERT INTO curriculum "
                    f"VALUES ({year}, 10, '{direction}', {sub_direction}, '{subject}', '{level}', "
                    f"{half_10_1}, {weeks_10_1}, '{report_10_1}', {half_10_2}, {weeks_10_2}, '{report_10_2}', "
                    f"{n_year_10})"
                )
                cur.connection.commit()

                half_11_1, report_11_1, weeks_11_1, half_11_2, report_11_2, weeks_11_2 = burden['111'][0], \
                                                                                         burden['111'][1], \
                                                                                         burden['111'][2], \
                                                                                         burden['112'][0], \
                                                                                         burden['112'][1], \
                                                                                         burden['112'][2]
                half_11_1 = half_11_1 if not (half_11_1 is None) else 0
                half_11_2 = half_11_2 if not (half_11_2 is None) else 0

                n_year_11 = weeks_11_1 * half_11_1 + weeks_11_2 * half_11_2

                report_11_1 = report_11_1 if not (report_11_1 is None) else 'NULL'
                report_11_2 = report_11_2 if not (report_11_2 is None) else 'NULL'
                cur.execute(
                    f"INSERT INTO curriculum "
                    f"VALUES ({year}, 11, '{direction}', {sub_direction}, '{subject}', '{level}', "
                    f"{half_11_1}, {weeks_11_1}, '{report_11_1}', {half_11_2}, {weeks_11_2}, '{report_11_2}', "
                    f"{n_year_11})"
                )
                cur.connection.commit()


def import_foreign_languages(con, cur):
    """Функция import_foreign_languages отвечает за импорт данных о языковых группах в базу данных."""
    with open('tables/foreign_languages.csv', encoding='utf-8') as r_file:
        data = csv.DictReader(r_file, delimiter=';')
        for row in data:
            print(row)
            x = "\ufefflang"
            teacher = row['teacher']
            cur.execute(
                f"SELECT id FROM teachers WHERE name='{teacher}'"
            )
            id_teacher = cur.fetchone()[0]
            cur.execute(
                f"INSERT INTO foreign_languages "
                f"VALUES (NULL, '{row[x]}', {int(row['n_group'])}, '{row['language']}', {id_teacher}, "
                f"'{row['teacher']}')"
            )
            con.commit()


def import_kvds_new(con, cur):
    wb = load_workbook(filename='tables/kvd.xlsx')
    sheet = wb.sheetnames[0]
    result = dict()
    temp = wb[sheet]['F1:CA2']
    for i in range(len(temp[0])):
        if temp[0][i].value in result:
            for elem in temp[1][i].value.split('/'):
                if not (elem in result[temp[0][i].value][0]):
                    result[temp[0][i].value][0].append(elem)
        else:
            result[temp[0][i].value] = [temp[1][i].value.split('/'), []]

    i = 3
    while not (wb[sheet][f'A{i}'].value is None):
        print(wb[sheet][f'B{i}'].value)
        student = wb[sheet][f'B{i}'].value.strip()
        cur.execute(
            f"SELECT student_id FROM students WHERE fullname='{student}'"
        )
        id_student = cur.fetchone()[0]
        courses = wb[sheet]['F1:CA1'][0]
        data = wb[sheet][f'F{i}:CA{i}'][0]
        for k in range(len(data)):
            if data[k].value == 1:
                result[courses[k].value][1].append(id_student)
        i += 1

    for elem in result:
        print(elem, result[elem])


def import_kvds(con, cur):
    """Функция import_kvds загружает информацию о курсах внеурочной деятельности гимназистов.
    В начале формируется словарь, в котором ключ - название курса, значение - список id слушателей.
    Первый элемент списка слушателей - преподаватель курса"""

    # Словарь temp
    temp = dict()
    with open('tables/kvd.csv', encoding='utf-8') as r_file:
        data = csv.DictReader(r_file, delimiter=';')
        count = 0
        for row in data:
            if count == 0:
                for item in row:
                    # формируем ключи и указываем первым элементом словаря - ведущего преподавателя
                    if row[item] != 'tmp' and len(item) > 1:
                        temp[item] = [row[item]]
            else:
                for item in row:
                    if item in temp and row[item] == '1':
                        student = row['ФИО ученика'].split()
                        if len(student) < 3:
                            continue
                        cur.execute(
                            f"SELECT student_id FROM students WHERE surname='{student[0]}' and name='{student[1]}'"
                        )
                        id_student = cur.fetchone()
                        if not (id_student is None):
                            id_student = id_student[0]
                        temp[item].append(id_student)
            count += 1

    for item in temp:
        # случай, когда модуль ведут несколько преподавателей
        for elem in temp[item][0].split('/'):
            cur.execute(
                f"INSERT INTO kvds (name_kvd, teacher, students_id) "
                f"VALUES ('{item}', '{elem}', '{';'.join(map(str, temp[item][1:]))}')"
            )
            con.commit()

    # update_kvds_in_students(con, cur)


def import_projects(con, cur):
    """Функция импорта проектов, присвоения им номера, а также записи номера проекта в таблицу student"""
    wb = load_workbook(filename='tables/projects.xlsx')
    sheets = wb.sheetnames
    temp = []
    for sheet in sheets:
        i = 2
        while not (wb[sheet][f'A{i}'].value is None):
            row = wb[sheet][f'A{i}:D{i}'][0]
            student = row[1].value
            project = row[3].value.strip()

            if not (project in temp):
                temp.append(project)
                cur.execute(
                    "INSERT INTO projects "
                    f"VALUES (NULL, '{project.strip()}', 0, 'NULL')"
                )

            id_project = temp.index(project) + 1
            id_student = functions.get_student_id(cur, student)
            print(student)
            functions.enrollment_project(con, cur, id_project, id_student)
            i += 1
    wb.close()


def import_students(con, cur):
    """Функция import_students добавляет в БД контингент гимназистов. Данные берутся из таблицы students.csv"""
    lng = {
        'Английский язык': 'en', 'Немецкий язык': 'ge',
        'Французский язык': 'fr', 'Китайский язык': 'ch',
        'Испанский язык': 'sp', 'Итальянский язык': 'it'
    }
    # читаем таблицу
    with open('tables/students.csv', encoding='utf-8') as file:
        data = csv.DictReader(file, delimiter=';')

        for row in data:
            fullname = row['ФИО'].strip()
            surname, name, last_name = row['ФИО'].split()
            tmp = row['Подпрофиль'].split()
            course = int(tmp[0])

            year = 2020 if course == 10 else 2019
            # отлавливаем 8 и 9 классы
            if course > 9:
                print(tmp)
                direction, sub_direction = tmp[1].split('-')[0], int(tmp[1][-1])
                print(direction, sub_direction)
                cur.execute(
                    f"SELECT id FROM config WHERE abb='{direction}' and id_subdirection={sub_direction} and year={year}"
                )
                direct = cur.fetchall()[0][0]
                print(direct)
            else:
                direct = 0

            # информация о языковых группах
            sec = lng[row['Второй ин. язык']]
            languages = f"en{row['Подгруппа']}{sec}{row['гр.']}"

            cur.execute(
                f"INSERT INTO students "
                f"VALUES (NULL, '{fullname}', '{surname}', '{name}', '{last_name}', {course}, '{direct}', "
                f"'{languages}', '', '')"
            )

            con.commit()


def import_kvds_new_new(con, cur):
    wb = load_workbook(filename='tables/kvd.xlsx')
    sheet = wb.sheetnames[0]
    kvds = dict()
    i = ""
    j = "F"
    while not (wb[sheet][f"{i}{j}1"].value is None):
        name = wb[sheet][f"{i}{j}1"].value
        tmp = wb[sheet][f"{i}{j}2"].value.split('/')
        teachers = list()
        for teacher in tmp:
            cur.execute(
                f"SELECT id FROM teachers WHERE name='{teacher}'"
            )
            teachers.append(cur.fetchone()[0])
        students = list()
        k = 3
        while not (wb[sheet][f"B{k}"].value is None):
            if wb[sheet][f"{i}{j}{k}"].value == 1:
                cur.execute(
                    f"SELECT student_id FROM students WHERE fullname='{wb[sheet][f'B{k}'].value}'"
                )
                student = cur.fetchone()[0]
                students.append(student)
                # students.add(wb[sheet][f"B{k}"].value)
            k += 1
        kvds[name] = {'teachers': teachers, 'students': students}
        j = chr(ord(j) + 1)
        if ord(j) > ord("Z"):
            j = "A"
            i = "A" if i == "" else chr(ord(i) + 1)
    # print(kvds)
    for kvd in kvds:
        print(kvd)
        for item in kvds[kvd]:
            print(item, ": ", kvds[kvd][item])
    for kvd in kvds:
        name = kvd
        students = ";".join(map(str, kvds[name]['students']))
        for teacher_id in kvds[name]['teachers']:
            cur.execute(
                f"SELECT name FROM teachers WHERE id='{teacher_id}'"
            )
            teacher = cur.fetchone()[0]
            print(name, teacher_id, teacher, students)
            cur.execute(
                f"INSERT INTO kvds "
                f"VALUES (NULL, '{name}', {teacher_id}, '{teacher}', '{students}')"
            )
            con.commit()


def update_students_kvds(con, cur):
    wb = load_workbook(filename='tables/kvd.xlsx')
    sheet = wb.sheetnames[0]
    t = wb[sheet]
    i = 3
    while not (t[f"B{i}"].value is None):
        m = ""
        n = "F"
        kvds = list()
        while not (t[f"{m}{n}1"].value is None):
            # print(1)
            if t[f"{m}{n}{i}"].value == 1:
                kvd = t[f"{m}{n}1"].value
                cur.execute(
                    f"SELECT id_kvd FROM kvds WHERE name_kvd='{kvd}'"
                )
                kvd_id = cur.fetchone()[0]
                cur.fetchall()
                kvds.append(kvd_id)
            n = chr(ord(n) + 1)
            if ord(n) > ord("Z"):
                n = "A"
                m = "A" if m == "" else chr(ord(m) + 1)
        # print(kvds)
        kvds = ";".join(map(str, kvds))
        print(t[f"B{i}"].value, kvds)
        cur.execute(
            f"UPDATE students SET kvds='{kvds}' WHERE fullname='{t[f'B{i}'].value}'"
        )
        con.commit()
        i += 1


def import_subjects(con, cur):
    """Функция import_subject загружает в БД информацию о всех предметах, кроме иностранных языков, включенных
    в расписание урок. Данные берутся из таблицу с расписанием. При формировании таблицы необходимо указывать номер
    подпрофиля, особенно для профильных предметов. Общие предметы для класса обозначаются цифрой 0."""
    with open('tables/subjects.csv', encoding='utf-8') as r_file:
        data = csv.DictReader(r_file, delimiter=';')
        for row in data:
            subject = row['\ufeffПредмет']

            teacher = row['Преподаватель']
            # узнаем id учителя
            cur.execute(
                f"SELECT id FROM teachers WHERE name='{teacher}'"
            )
            id_teacher = cur.fetchone()[0]

            course = int(row['Класс'])
            year = 2020 if course == 10 else 2019
            direction = row['Профиль']
            sub_direction = int(row['Подпрофиль'])
            if sub_direction == 0:
                # случай, когда предмет у всего класса, то есть у нескольких профилей
                cur.execute(
                    f"SELECT id FROM config WHERE abb='{direction}' and year={year}"
                )
                ids = cur.fetchall()
                for elem in ids:
                    cur.execute(
                        "INSERT INTO subjects "
                        f"VALUES (NULL, '{subject}', {id_teacher}, {elem[0]})"
                    )
                    con.commit()
            else:
                cur.execute(
                    f"SELECT id FROM config WHERE abb='{direction}' and year={year} and id_subdirection={sub_direction}"
                )
                id = cur.fetchone()[0]
                cur.execute(
                    "INSERT INTO subjects "
                    f"VALUES (NULL, '{subject}', {id_teacher}, {id})"
                )
                con.commit()

            # cur.execute(
            #    f"INSERT INTO subjects "
            #    f"VALUES (NULL, '{subject}', NULL, '{teacher}', {course}, '{direction}', {sub_direction})"
            # )
            # con.commit()


def parse_excel_to_dict(files):
    """Функция parse_excel_to_dict получает в качестве аргумента список имен файлов, содержащих учебные планы.
    На данный момент вручную установлены значения этого списка.
    На выходе словарь:
        ключ - направление подготовки по формату <класс><профиль>-<подпрофиль>-<год>,
        значение - словарь:
            ключ - учебный предмет, изучаемый в профиле,
            значение - картеж из двух элементов:
                первый - уровень изучения (базовый/углубленный),
                второй - словарь"""
    result = dict()

    for name in files:
        # подключаемся к excel таблице
        wb = load_workbook(filename=name)
        # получаем список листов в книге
        sheets = wb.sheetnames
        # получаем год, который загружается
        suffix = name.split('.')[0][-4:]
        # обходим листы
        for sheet in sheets:
            direction = sheet + f'-{suffix}'
            result[direction] = []
            i = 16
            count = 0
            while count <= 2:
                # нужно пропустить строки 33 и 34, так как они пустые
                if wb[sheet][f'B{i}'].value is None:
                    count += 1
                    i += 1
                    continue
                # обработка обязательной части ООП
                data = wb[sheet][f'B{i}:L{i}'][0]
                subject = data[0].value.rstrip('*') if i < 33 else f'**{data[0].value}'.rstrip('*')
                level = data[1].value
                burden = {
                    '101': (data[2].value, data[3].value, 16),
                    '102': (data[4].value, data[5].value, 19),
                    '111': (data[6].value, data[7].value, 16),
                    '112': (data[8].value, data[9].value, 19)
                }
                temp = dict()
                temp[subject] = (level, burden)
                result[direction].append(temp)
                i += 1
        wb.close()
    return result


def update_kvds_in_students(con, cur):
    """Функция upload_kvds_in_students обновляет данные в таблице со контингентом (students) информацию о курсах КВД,
    которые они посещают. Необходимые курсы записываются в столбик kvds через точку с запятой"""

    # извлекаем информацию о курсах
    cur.execute(
        "SELECT * FROM kvds"
    )

    data = cur.fetchall()

    # обходим курсы и смотрим, какие студенты их посещают
    for item in data:
        for id_student in str(item[4]).split(';')[:-1]:
            # узнаем список курсов конкретного студента
            cur.execute(
                f"SELECT kvds FROM students WHERE student_id='{id_student}'"
            )
            students_kvds = cur.fetchone()
            if not (students_kvds is None):
                students_kvds = students_kvds[0]

            if students_kvds is None:
                cur.execute(
                    f"UPDATE students SET kvds='{str(item[0])}' WHERE student_id={int(id_student)}"
                )
                con.commit()
            else:
                # если такой курс уже есть в списке слушателя, то его добавлять не нужно
                if not (str(item[0]) in str(students_kvds)):
                    students_kvds = str(students_kvds) + ';' + str(item[0])
                    cur.execute(
                        f"UPDATE students SET kvds='{students_kvds}' WHERE student_id={int(id_student)}"
                    )
                    con.commit()


def import_directions(con, cur, files):
    """Функция отвечает за импорт данных о направлениях подготовки, взятые из файлов с учебными планами."""
    data = dict()
    for file_name in files:
        year = file_name.split('.')[0].split('/')[1][3:]
        data[year] = dict()
        wb = load_workbook(filename=file_name)
        sheets = wb.sheetnames
        for sheet in sheets:
            id_sub_direction = int(sheet[-1])
            direction = wb[sheet]['a9'].value.split()[
                wb[sheet]['a9'].value.split().index('профиля') - 1
                ].rstrip(')').lstrip('(')
            sub_direction = wb[sheet]['a10'].value.split('«')[1].rstrip('»').lower()
            temp = direction.split('-')
            if len(temp) == 1:
                abb = temp[0][:3].upper()
            else:
                abb = temp[0][0].upper() + temp[1][0].upper()

            if direction in data[year]:
                data[year][direction].append((id_sub_direction, sub_direction, abb))
            else:
                data[year][direction] = [(id_sub_direction, sub_direction, abb)]

    for year in data:
        k = 0
        for direction in data[year]:
            k += 1
            for item in data[year][direction]:
                # уникальный код направления: <год><номер профиля (2 цифры)><номер подпрофиля (2 цифры)>
                uni_id = int(f'{year}{k // 10}{k % 10}{item[0] // 10}{item[0] % 10}')
                cur.execute(
                    f"INSERT INTO config "
                    f"VALUES ({uni_id}, {year}, {k}, '{direction}', '{item[0]}', '{item[1]}', '{item[2]}')"
                )
                con.commit()


def engine():
    con = sqlite3.connect('telebot_db.db')
    cur = con.cursor()
    # import_students(cur)
    # import_kvds(cur)
    # import_subjects(cur)
    # import_curriculum(cur, parse_excel_to_dict(['tables/cur2020.xlsx', 'tables/cur2019.xlsx']))
    # import_foreign_languages(cur)
    # import_projects(cur)
    # import_directions(cur, ['tables/cur2020.xlsx', 'tables/cur2019.xlsx'])
    con.close()


engine()
